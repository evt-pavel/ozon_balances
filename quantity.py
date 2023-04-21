from openpyxl import load_workbook
from datetime import date, timedelta
from send_email import send_email


tride = load_workbook(filename='tride.xlsx')
ozon = load_workbook(filename='ozon.xlsx')


tride_sheets = tride.sheetnames
ozon_sheets = ozon.sheetnames

tride_sheet = tride[tride_sheets[0]]
ozon_sheet = ozon[ozon_sheets[1]]

today = date.today().strftime("%d-%b-%Y")
yesterday = (date.today() - timedelta(days=1)).strftime("%d-%b-%Y")

    # проверка на пустые строки
def type_object(word):
    if word is None:
        return word
    else:
        return word.replace(' ', '')

# приводим все занчения из столбца количество к 0 в остатках озона
def valueToZero():
    for i in range(1, ozon_sheet.max_row):
        if type(ozon_sheet.cell(row=i, column=5).value) is str:
            continue
        else:
            ozon_sheet.cell(row=i, column=5).value = 0


# соотносим артикли в прайсе поставщика с остатками озона
def articles():
    count = 0
    for i in range(1, ozon_sheet.max_row):
        x = type_object(ozon_sheet.cell(row=i, column=3).value)

        for j in range(1, tride_sheet.max_row):
            y = type_object(tride_sheet.cell(row=j, column=3).value)

            if y == x and ozon_sheet.cell(row=i, column=2).value is None:
                count += 1
                ozon_sheet.cell(row=i, column=2).value = tride_sheet.cell(row=j, column=1).value
                print(x, y)
    print(f'Добавлено {count} новых артикулов')


 # проверка остатков
def quantity():
    count = 0
    
    for i in range(1, ozon_sheet.max_row):
        x = type_object(ozon_sheet.cell(row=i, column=2).value)

        for j in range(1, tride_sheet.max_row):
            y = type_object(tride_sheet.cell(row=j, column=1).value)


            if x == y:
                count += 1

                if tride_sheet.cell(row=j, column=4).value == 'Резерв':
                    ozon_sheet.cell(row=i, column=5).value = 0
                elif tride_sheet.cell(row=j, column=4).value == '2+':
                    ozon_sheet.cell(row=i, column=5).value = 15
                else:
                    if tride_sheet.cell(row=j, column=4).value == 'Склад':
                        continue
                    else:
                        ozon_sheet.cell(row=i, column=5).value = 12

    ozon_sheet.delete_cols(2)
    print(f'Обновлено {count} товаров!')
    ozon.save(f'ozon_{today}.xlsx')
    ozon.close()


def value_comparison():
    

    quantity_yesterday = load_workbook(filename=f'ozon_{yesterday}.xlsx')
    quantity_today = load_workbook(filename=f'ozon_{today}.xlsx')

    yesterday_sheets = quantity_yesterday.sheetnames
    today_sheets = quantity_today.sheetnames

    quantity_yesterday_sheet = quantity_yesterday[yesterday_sheets[1]] 
    quantity_today_sheet = quantity_today[today_sheets[1]]
    x = 0
    y = False

    while y is False:
        y = True

        for i in range(2, quantity_today_sheet.max_row):

            if quantity_today_sheet.cell(row=i, column=4).value == quantity_yesterday_sheet.cell(row=i, column=4).value\
                and quantity_today_sheet.cell(row=i, column=4).value != None and quantity_yesterday_sheet.cell(row=i, column=4).value != None:
                quantity_today_sheet.delete_rows(i)
                quantity_yesterday_sheet.delete_rows(i)
                x += 1
                y = False

        
    quantity_today.save(f'comparison_{today}.xlsx')
    return print(x, 'совпадений!')

